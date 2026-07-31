"""工程设计包服务：把工业品设计任务产物打包为可交付的工程包（zip + PDF + XLSX）。"""
from __future__ import annotations

import io
import logging
import re
import tempfile
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Mapping
from uuid import UUID

from sqlalchemy.orm import Session

from app.config.settings import settings
from app.services.asset_blob_service import AssetBlobService
from app.services.workflow_task_repository import WorkflowTaskRepository

logger = logging.getLogger(__name__)

JsonValue = None | bool | int | float | str | list["JsonValue"] | dict[str, "JsonValue"]

ASSET_KEYS: list[tuple[str, str, str]] = [
    ("planLineSvgAssetId", "方案图.svg", "svg"),
    ("planLineDxfAssetId", "工程图.dxf", "dxf"),
    ("renderPngAssetId", "设计图.png", "png"),
    ("enhancedImageAssetId", "精修图.png", "png"),
    ("modelStepAssetId", "三维模型.step", "step"),
    ("modelStlAssetId", "三维模型.stl", "stl"),
    ("modelGlbAssetId", "三维预览.glb", "glb"),
    ("modelScriptAssetId", "三维模型.py", "py"),
]

_PDF_TITLE_FONT = "NotoSansCJK"
_PDF_BODY_FONT = "NotoSansCJK"


class EngineeringPackageServiceError(Exception):
    """工程设计包服务可预期错误。"""

    def __init__(self, message: str, error_code: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.message = message
        self.error_code = error_code
        self.status_code = status_code


class EngineeringPackageService:
    """负责把任务 outputs 关联资产打包成工程包。"""

    def __init__(
        self,
        *,
        asset_service: AssetBlobService | None = None,
        repository_factory=WorkflowTaskRepository,
        runtime_temp_root: Path | None = None,
    ) -> None:
        self.asset_service = asset_service or AssetBlobService(
            chunk_size=settings.ASSET_CHUNK_SIZE_BYTES
        )
        self.repository_factory = repository_factory
        self.runtime_temp_root = runtime_temp_root

    @staticmethod
    def _parse_asset_uuid(value: object) -> UUID | None:
        if not isinstance(value, str):
            return None
        try:
            return UUID(value)
        except ValueError:
            return None

    @staticmethod
    def _extract_asset_id(value: object) -> str | None:
        if not isinstance(value, str):
            return None
        if EngineeringPackageService._parse_asset_uuid(value) is not None:
            return value
        match = re.match(r".*?/([0-9a-fA-F-]{36})/download", value)
        return match.group(1) if match else None

    def _read_asset(self, db: Session, user_id: str, asset_id: str) -> bytes:
        try:
            return self.asset_service.read_bytes(db, UUID(asset_id), user_id)
        except Exception as exc:
            raise EngineeringPackageServiceError(
                f"读取资产失败：{asset_id}",
                "PACKAGE_ASSET_READ_FAILED",
                status_code=502,
            ) from exc

    def _collect_assets(
        self,
        db: Session,
        user_id: str,
        outputs: Mapping[str, object],
    ) -> dict[str, bytes]:
        collected: dict[str, bytes] = {}
        for key, filename, _kind in ASSET_KEYS:
            asset_id = self._extract_asset_id(outputs.get(key))
            if asset_id is None:
                continue
            try:
                collected[filename] = self._read_asset(db, user_id, asset_id)
            except EngineeringPackageServiceError:
                logger.warning("跳过无法读取的资产 %s", filename)
        render_views = outputs.get("renderViews")
        if isinstance(render_views, list):
            for index, item in enumerate(render_views):
                if not isinstance(item, Mapping):
                    continue
                render_id = self._extract_asset_id(item.get("assetId"))
                if render_id is None:
                    continue
                key = item.get("key")
                filename = f"渲染图/{key or f'view_{index}'}.png"
                try:
                    collected[filename] = self._read_asset(db, user_id, render_id)
                except EngineeringPackageServiceError:
                    logger.warning("跳过无法读取的渲染图 %s", filename)
        return collected

    def _build_pdf(
        self,
        *,
        project_name: str,
        design_spec: Mapping[str, object] | None,
        input_payload: Mapping[str, object] | None,
        collected: Mapping[str, bytes],
    ) -> bytes:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.platypus import (
                Image,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
            )
        except ImportError as exc:
            raise EngineeringPackageServiceError(
                "报告生成组件不可用：缺少 reportlab",
                "PACKAGE_PDF_DEP_UNAVAILABLE",
                status_code=503,
            ) from exc

        font_path = Path("/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc")
        try:
            if _PDF_TITLE_FONT not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(_PDF_TITLE_FONT, str(font_path)))
        except Exception as exc:
            raise EngineeringPackageServiceError(
                "报告生成失败：缺少系统中文字体（fonts-noto-cjk）",
                "PACKAGE_PDF_FONT_UNAVAILABLE",
                status_code=503,
            ) from exc

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=18 * mm,
            leftMargin=18 * mm,
            topMargin=16 * mm,
            bottomMargin=16 * mm,
        )
        title_style = ParagraphStyle(
            "title", fontName=_PDF_TITLE_FONT, fontSize=20, leading=26, spaceAfter=12
        )
        heading_style = ParagraphStyle(
            "heading", fontName=_PDF_TITLE_FONT, fontSize=13, leading=18, spaceBefore=8, spaceAfter=4
        )
        body_style = ParagraphStyle(
            "body", fontName=_PDF_TITLE_FONT, fontSize=9.5, leading=14, spaceAfter=3
        )

        design_spec = design_spec or {}
        requirement = design_spec.get("requirementText") or input_payload.get("text") or ""
        industry = design_spec.get("industry") or "装备制造"
        content_type = design_spec.get("inputType") or (input_payload.get("inputType") if input_payload else None) or "text"

        story: list[object] = [
            Paragraph("工业设计工程包 · 设计说明", title_style),
            Paragraph(f"项目名称：{project_name}", body_style),
            Paragraph(f"所属行业：{industry}", body_style),
            Paragraph(f"输入类型：{content_type}", body_style),
            Spacer(1, 6),
            Paragraph("一、需求描述", heading_style),
            Paragraph(str(requirement or "未提供设计描述。"), body_style),
            Spacer(1, 6),
            Paragraph("二、产物清单", heading_style),
        ]
        product_lines = {
            "方案图.svg": "2D 平面图（CAD 线稿）",
            "工程图.dxf": "2D 工程图（CAD 线稿）",
            "设计图.png": "AI 渲染效果图",
            "精修图.png": "商业精修图",
            "三维模型.step": "3D 数模（STEP）",
            "三维模型.stl": "3D 网格（STL）",
            "三维预览.glb": "3D 预览（GLB）",
            "三维模型.py": "参数化建模脚本",
        }
        for filename, desc in product_lines.items():
            if filename in collected:
                story.append(Paragraph(f"• {filename} —— {desc}", body_style))
        render_count = sum(1 for key in collected if key.startswith("渲染图/"))
        if render_count:
            story.append(Paragraph(f"• 渲染图/ —— 多视角 × 配色渲染图（{render_count} 张）", body_style))

        preview = collected.get("设计图.png")
        if preview is None:
            for key, value in collected.items():
                if key.startswith("渲染图/"):
                    preview = value
                    break
        if preview is not None:
            story.append(Spacer(1, 8))
            story.append(Paragraph("三、预览图", heading_style))
            try:
                from PIL import Image as PILImage

                img = PILImage.open(io.BytesIO(preview))
                img.thumbnail((560, 420))
                preview_buffer = io.BytesIO()
                img.save(preview_buffer, format="PNG")
                story.append(Image(preview_buffer, width=140 * mm, height=105 * mm))
            except Exception:
                logger.warning("PDF 内嵌预览图失败")

        story.append(Spacer(1, 8))
        story.append(
            Paragraph(
                f"生成时间：{datetime.now(timezone.utc).astimezone().strftime('%Y-%m-%d %H:%M:%S')}",
                body_style,
            )
        )
        doc.build(story)
        return buffer.getvalue()

    def _build_bom_xlsx(
        self,
        *,
        project_name: str,
        design_spec: Mapping[str, object] | None,
        collected: Mapping[str, bytes],
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise EngineeringPackageServiceError(
                "BOM 生成组件不可用：缺少 openpyxl",
                "PACKAGE_XLSX_DEP_UNAVAILABLE",
                status_code=503,
            ) from exc

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"
        ws.append(["序号", "名称", "材料", "数量", "规格说明", "备注"])
        header_fill = PatternFill(start_color="DCE9F7", end_color="DCE9F7", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        ws.append([1, project_name, "待确认（建议 6061 铝 / ABS / 45 钢）", 1, "整体件", "按 3D 数模尺寸制造"])
        ws.append([2, "紧固件", "标准件", "按图纸", "安装孔位", "以 2D 工程图孔位为准"])
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(40, max_length + 4)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def build_package(
        self,
        *,
        db: Session,
        user_id: str,
        task_id: str,
        publish_assets: bool = True,
    ) -> dict[str, JsonValue]:
        """按任务 ID 组装工程设计包并入库，返回 zip 资产下载信息。"""
        task = self.repository_factory(db).get_internal(task_id)
        if task is None:
            raise EngineeringPackageServiceError(
                "工业品设计任务不存在",
                "PACKAGE_TASK_NOT_FOUND",
                status_code=404,
            )
        outputs = task.get("outputs") or {}
        if not isinstance(outputs, Mapping):
            raise EngineeringPackageServiceError(
                "任务产物数据异常",
                "PACKAGE_OUTPUTS_INVALID",
                status_code=500,
            )
        if not outputs.get("modelStepAssetId") and not outputs.get("modelStlAssetId"):
            raise EngineeringPackageServiceError(
                "该任务尚未产出 3D 模型，无法生成工程包",
                "PACKAGE_NO_MODEL",
                status_code=400,
            )
        project_name = str(
            task.get("projectId") or task.get("designSpec", {}).get("projectName") or "未命名项目"
        )
        design_spec = task.get("designSpec")
        input_payload = task.get("inputPayload")

        collected = self._collect_assets(db, user_id, outputs)
        pdf_bytes = self._build_pdf(
            project_name=project_name,
            design_spec=design_spec if isinstance(design_spec, Mapping) else None,
            input_payload=input_payload if isinstance(input_payload, Mapping) else None,
            collected=collected,
        )
        xlsx_bytes = self._build_bom_xlsx(
            project_name=project_name,
            design_spec=design_spec if isinstance(design_spec, Mapping) else None,
            collected=collected,
        )

        package_task_id = f"engineering_package_{uuid.uuid4().hex[:16]}"
        with tempfile.TemporaryDirectory(
            prefix="engineering-package-",
            dir=self.runtime_temp_root,
        ) as temporary_directory:
            work_dir = Path(temporary_directory)
            zip_path = work_dir / "工程包.zip"
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as archive:
                archive.writestr("设计说明.pdf", pdf_bytes)
                archive.writestr("BOM.xlsx", xlsx_bytes)
                for filename, content in collected.items():
                    archive.writestr(filename, content)

            asset = self.asset_service.store_bytes(
                db=db,
                user_id=user_id,
                filename=f"{package_task_id}.zip",
                content_type="application/zip",
                kind="archive",
                source="generated",
                content=zip_path.read_bytes(),
                task_id=task_id,
                metadata={
                    "engineeringPackageTaskId": package_task_id,
                    "format": "zip",
                },
                publish=publish_assets,
            )
        return {
            "taskId": package_task_id,
            "status": "completed",
            "packageAssetId": str(asset.id),
            "packageDownloadUrl": f"{settings.API_V1_PREFIX}/assets/{asset.id}/download",
            "filename": "工程包.zip",
        }


engineering_package_service = EngineeringPackageService()
